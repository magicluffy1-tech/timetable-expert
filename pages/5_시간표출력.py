"""
pages/5_시간표출력.py
학급별/교사별/특별실별 시간표 인쇄용 HTML 출력 및 Excel 다운로드 지원
"""
import streamlit as st
import pandas as pd
import io
import sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from utils.data_manager import (
    get_data, get_all_classes, get_subject_color,
    get_teacher_by_id, get_teachers_for_subject_grade,
    get_periods_for_day
)
from utils.constraints import Assignment

st.set_page_config(page_title="시간표출력 | 시간표 전문가", page_icon="🖨️", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght=300;400;500;700;900&display=swap');
html, body, [class*="css"] { font-family: 'Noto Sans KR', sans-serif; }
.stApp { background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%); }
.page-header {
    background: linear-gradient(135deg, #36d1dc 0%, #5b86e5 100%);
    border-radius: 16px; padding: 28px 36px; margin-bottom: 28px;
    box-shadow: 0 12px 40px rgba(91,134,229,0.35);
}
.page-header h2 { color: white; font-size: 1.8rem; font-weight: 900; margin: 0 0 6px 0; }
.page-header p  { color: rgba(255,255,255,0.8); margin: 0; }
.card {
    background: rgba(255,255,255,0.06); backdrop-filter: blur(12px);
    border: 1px solid rgba(255,255,255,0.12); border-radius: 14px; padding: 20px; margin-bottom: 12px;
}
.card-title {
    font-size: 0.95rem; font-weight: 700; color: rgba(255,255,255,0.9);
    border-left: 4px solid #5b86e5; padding-left: 10px; margin-bottom: 14px;
}
.stButton > button {
    background: linear-gradient(135deg, #36d1dc, #5b86e5);
    color: white; border: none; border-radius: 10px; font-weight: 700;
    font-family: 'Noto Sans KR', sans-serif;
}

/* 인쇄 스타일 */
.print-area {
    background: white; color: black; border-radius: 8px; padding: 30px; margin-top: 10px;
    box-shadow: 0 10px 30px rgba(0,0,0,0.5);
}
.print-title {
    font-size: 1.8rem; font-weight: 800; text-align: center; margin-bottom: 20px; color: #333;
}
.print-table {
    width: 100%; border-collapse: collapse; text-align: center;
}
.print-table th, .print-table td {
    border: 1px solid #333; padding: 8px; font-size: 0.9rem;
}
.print-table th { background: #f2f2f2; font-weight: bold; }
.print-table td { height: 60px; vertical-align: middle; }
.print-subj { font-weight: bold; font-size: 1rem; }
.print-teach { font-size: 0.8rem; color: #555; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="page-header">
    <h2>🖨️ 시간표 출력 및 다운로드</h2>
    <p>완성된 시간표를 학급별, 교사별, 특별실별로 인쇄용 HTML로 보거나 Excel 파일로 다운로드합니다.</p>
</div>
""", unsafe_allow_html=True)

data = get_data()
all_classes = get_all_classes(data)
days = data["school_info"].get("days", ["월","화","수","목","금"])
periods_n = data["school_info"].get("periods_per_day", 7)
teachers = data.get("teachers", [])
special_rooms = data.get("special_rooms", [])

if not all_classes:
    st.warning("학급 정보가 없습니다. 학교기본설정에서 학급을 먼저 구성하세요.")
    st.stop()

# 시간표 데이터 유무 체크
timetable = data.get("timetable", {})
has_timetable = any(timetable.get(cls) for cls in all_classes)

if not has_timetable:
    st.info("⚠️ 생성되어 확정된 시간표가 없습니다. [시간표편성] 페이지에서 자동 편성을 먼저 실행하세요.")
    st.stop()

tab1, tab2, tab3, tab4 = st.tabs(["🏫 학급별 출력", "👩‍🏫 교사별 출력", "🔬 특별실 배정표", "📥 Excel 내보내기"])

# ──────────────────────────────────────────────────────────────────────────────
# TAB 1: 학급별 출력
# ──────────────────────────────────────────────────────────────────────────────
with tab1:
    col_sel, col_print = st.columns([1, 4])
    with col_sel:
        st.markdown("**출력 대상**")
        sel_class = st.selectbox("학급 선택", all_classes, key="print_class_sel")
    
    with col_print:
        st.markdown('<div class="card-title">📄 인쇄 미리보기</div>', unsafe_allow_html=True)
        class_tt = timetable.get(sel_class, {})
        
        # HTML 인쇄용 테이블 생성
        html = f"""
        <div class="print-area">
            <div class="print-title">{data['school_info'].get('name','○○중학교')} - {sel_class} 주간 시간표</div>
            <table class="print-table">
                <thead>
                    <tr>
                        <th style="width: 10%;">교시</th>
        """
        for d in days:
            html += f"<th style='width: 18%;'>{d}요일</th>"
        html += "</tr></thead><tbody>"
        
        max_periods = max(get_periods_for_day(data, d) for d in days) if days else 7
        for p in range(1, max_periods + 1):
            html += f"<tr><td><strong>{p}교시</strong></td>"
            for d in days:
                periods_for_day = get_periods_for_day(data, d)
                if p > periods_for_day:
                    html += "<td><span style='color:#ccc; font-size:0.8rem;'>-</span></td>"
                    continue
                cell = class_tt.get(d, {}).get(str(p), {})
                subj = cell.get("subject", "")
                tid = cell.get("teacher_id", "")
                t_obj = get_teacher_by_id(data, tid)
                t_name = t_obj.get("name", "") if t_obj else ""
                
                if subj:
                    room_str = f" ({cell.get('special_room')})" if cell.get("special_room") else ""
                    html += f"<td><div class='print-subj'>{subj}{room_str}</div><div class='print-teach'>{t_name}</div></td>"
                else:
                    html += "<td></td>"
            html += "</tr>"
            
        html += "</tbody></table></div>"
        
        st.markdown(html, unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# TAB 2: 교사별 출력
# ──────────────────────────────────────────────────────────────────────────────
with tab2:
    if not teachers:
        st.info("등록된 교사가 없습니다.")
    else:
        col_t_sel, col_t_print = st.columns([1, 4])
        with col_t_sel:
            t_opts = {f"{t['name']} ({t['id']})": t for t in teachers}
            sel_t_label = st.selectbox("교사 선택", list(t_opts.keys()), key="print_teacher_sel")
            sel_teacher = t_opts[sel_t_label]
            
        with col_t_print:
            st.markdown('<div class="card-title">📄 인쇄 미리보기</div>', unsafe_allow_html=True)
            
            # 교사 시간표 재구성
            # 교사 시간표 재구성
            t_tt = {d: {p: None for p in range(1, get_periods_for_day(data, d) + 1)} for d in days}
            for cls in all_classes:
                cls_tt = timetable.get(cls, {})
                for d in days:
                    periods_for_day = get_periods_for_day(data, d)
                    for p in range(1, periods_for_day + 1):
                        cell = cls_tt.get(d, {}).get(str(p), {})
                        if cell.get("teacher_id") == sel_teacher["id"]:
                            t_tt[d][p] = {
                                "class_name": cls,
                                "subject": cell.get("subject", ""),
                                "room": cell.get("special_room", "")
                            }
            
            html = f"""
            <div class="print-area">
                <div class="print-title">{data['school_info'].get('name','○○중학교')} - {sel_teacher['name']} 선생님 주간 시간표</div>
                <table class="print-table">
                    <thead>
                        <tr>
                            <th style="width: 10%;">교시</th>
            """
            for d in days:
                html += f"<th style='width: 18%;'>{d}요일</th>"
            html += "</tr></thead><tbody>"
            
            max_periods = max(get_periods_for_day(data, d) for d in days) if days else 7
            for p in range(1, max_periods + 1):
                html += f"<tr><td><strong>{p}교시</strong></td>"
                for d in days:
                    periods_for_day = get_periods_for_day(data, d)
                    if p > periods_for_day:
                        html += "<td><span style='color:#ccc; font-size:0.8rem;'>-</span></td>"
                        continue
                    cell = t_tt[d].get(p)
                    if cell:
                        room_str = f" ({cell['room']})" if cell['room'] else ""
                        html += f"<td><div class='print-subj'>{cell['subject']}{room_str}</div><div class='print-teach'>{cell['class_name']}</div></td>"
                    else:
                        html += "<td><span style='color:#ccc; font-size:0.8rem;'>공강</span></td>"
                html += "</tr>"
                
            html += "</tbody></table></div>"
            st.markdown(html, unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# TAB 3: 특별실 배정표
# ──────────────────────────────────────────────────────────────────────────────
with tab3:
    if not special_rooms:
        st.info("등록된 특별실이 없습니다.")
    else:
        col_r_sel, col_r_print = st.columns([1, 4])
        with col_r_sel:
            r_opts = {r["name"]: r for r in special_rooms}
            sel_r_name = st.selectbox("특별실 선택", list(r_opts.keys()), key="print_room_sel")
            
        with col_r_print:
            st.markdown('<div class="card-title">📄 인쇄 미리보기</div>', unsafe_allow_html=True)
            
            # 특별실 시간표 재구성
            # 특별실 시간표 재구성
            r_tt = {d: {p: None for p in range(1, get_periods_for_day(data, d) + 1)} for d in days}
            for cls in all_classes:
                cls_tt = timetable.get(cls, {})
                for d in days:
                    periods_for_day = get_periods_for_day(data, d)
                    for p in range(1, periods_for_day + 1):
                        cell = cls_tt.get(d, {}).get(str(p), {})
                        if cell.get("special_room") == sel_r_name:
                            t_obj = get_teacher_by_id(data, cell.get("teacher_id"))
                            t_name = t_obj.get("name", "") if t_obj else ""
                            r_tt[d][p] = {
                                "class_name": cls,
                                "subject": cell.get("subject", ""),
                                "teacher": t_name
                            }
            
            html = f"""
            <div class="print-area">
                <div class="print-title">{data['school_info'].get('name','○○중학교')} - {sel_r_name} 주간 배정표</div>
                <table class="print-table">
                    <thead>
                        <tr>
                            <th style="width: 10%;">교시</th>
            """
            for d in days:
                html += f"<th style='width: 18%;'>{d}요일</th>"
            html += "</tr></thead><tbody>"
            
            max_periods = max(get_periods_for_day(data, d) for d in days) if days else 7
            for p in range(1, max_periods + 1):
                html += f"<tr><td><strong>{p}교시</strong></td>"
                for d in days:
                    periods_for_day = get_periods_for_day(data, d)
                    if p > periods_for_day:
                        html += "<td><span style='color:#ccc; font-size:0.8rem;'>-</span></td>"
                        continue
                    cell = r_tt[d].get(p)
                    if cell:
                        html += f"<td><div class='print-subj'>{cell['class_name']}</div><div class='print-teach'>{cell['subject']} ({cell['teacher']})</div></td>"
                    else:
                        html += "<td><span style='color:#ccc; font-size:0.8rem;'>미배정</span></td>"
                html += "</tr>"
                
            html += "</tbody></table></div>"
            st.markdown(html, unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# TAB 4: Excel 내보내기
# ──────────────────────────────────────────────────────────────────────────────
with tab4:
    st.markdown('<div class="card-title">📥 엑셀(Excel) 일괄 내보내기</div>', unsafe_allow_html=True)
    st.markdown("""
    모든 학급의 시간표가 포함된 하나의 깔끔한 엑셀 파일(.xlsx)을 다운로드할 수 있습니다.
    엑셀 파일은 각 학급별 탭으로 깔끔하게 자동 서식화되어 저장됩니다.
    """)
    
    if st.button("📥 고품질 Excel 파일 생성 및 다운로드", use_container_width=True):
        wb = Workbook()
        
        # 기본 Sheet 제거
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 얇은 테두리 스타일 지정
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # 각 학급별 시트 생성
        for cls in all_classes:
            ws = wb.create_sheet(title=cls)
            class_tt = timetable.get(cls, {})
            
            # 제목행 설정
            ws.merge_cells("A1:F1")
            title_cell = ws["A1"]
            title_cell.value = f"{data['school_info'].get('name','○○중학교')} {cls} 시간표"
            title_cell.font = Font(name="돋움", size=16, bold=True, color="1F497D")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40
            
            # 빈 행
            ws.row_dimensions[2].height = 10
            
            # 헤더 설정
            headers = ["교시"] + [f"{d}요일" for d in days]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(name="돋움", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[3].height = 25
            
            # 시간표 데이터 채우기
            max_periods = max(get_periods_for_day(data, d) for d in days) if days else 7
            for p in range(1, max_periods + 1):
                row_num = p + 3
                ws.row_dimensions[row_num].height = 35
                
                # 교시 열
                p_cell = ws.cell(row=row_num, column=1)
                p_cell.value = f"{p}교시"
                p_cell.font = Font(name="돋움", size=10, bold=True, color="333333")
                p_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                p_cell.alignment = Alignment(horizontal="center", vertical="center")
                p_cell.border = thin_border
                
                # 요일별 열
                for d_num, d in enumerate(days, 2):
                    cell = ws.cell(row=row_num, column=d_num)
                    periods_for_day = get_periods_for_day(data, d)
                    if p > periods_for_day:
                        cell.value = "-"
                        cell.font = Font(name="돋움", size=9, color="A0A0A0")
                        cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        continue

                    cell_data = class_tt.get(d, {}).get(str(p), {})
                    subj = cell_data.get("subject", "")
                    tid = cell_data.get("teacher_id", "")
                    t_obj = get_teacher_by_id(data, tid)
                    t_name = t_obj.get("name", "") if t_obj else ""
                    
                    if subj:
                        room_str = f" ({cell_data.get('special_room')})" if cell_data.get("special_room") else ""
                        cell.value = f"{subj}\n{t_name}{room_str}"
                        cell.font = Font(name="돋움", size=9)
                        
                        # 파스텔톤 배경색 연결
                        hex_color = get_subject_color(subj).replace("#", "")
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    else:
                        cell.value = ""
                    
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
            
            # 열 너비 조절
            ws.column_dimensions["A"].width = 10
            for col_letter in ["B", "C", "D", "E", "F"]:
                ws.column_dimensions[col_letter].width = 16
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.success("✅ 엑셀 파일이 성공적으로 생성되었습니다!")
        st.download_button(
            label="📥 엑셀 파일 다운로드",
            data=output,
            file_name=f"{data['school_info'].get('name','middle_school')}_timetable.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
